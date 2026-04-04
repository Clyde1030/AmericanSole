"""
Preprocessor — cleans a vendor shipment Excel file before LLM extraction.

The input Excel workbook contains two worksheets produced by upstream factories:
  1. "AS" — pending/in-progress shipments
  2. "Shipped" — completed shipments with container and arrival info

This script strips noise columns (photos, WIP status, internal dates) and
normalises merged cells so the downstream LLM agent receives a flat,
machine-readable table.

Processing steps per worksheet:

Both worksheets:
  - PO numbers are coerced to text strings. Excel may store them as numbers
    (e.g., 12345.0); this step converts them to clean strings ("12345") and
    sets the cell data type to "s" (string) so downstream consumers never
    misinterpret them as numeric values.

AS worksheet:
  1. Unmerge all header-row cells and forward-fill values.
  2. Coerce PO numbers to strings.
  3. Identify and delete columns:
     Photo, WIP (cutting/stitching/last), ETD ShenZhen, AS XF, Update AS XF.

Shipped worksheet:
  1. Unhide any hidden columns so all data is visible.
  2. Unmerge all header-row cells and forward-fill values.
  3. Unmerge and forward-fill data cells in grouped columns:
     LH XF, Container Type, ETD-ShenZhen, ETA-SA, Remark.
  4. Coerce PO numbers to strings.
  5. Filter rows by date: keep only rows whose "LH XF" (factory ex-factory
     date) falls on or after January 1 of the current year. Older shipments
     are removed to limit the dataset to the active shipping window.
  6. Identify and delete columns:
     Photo, WIP (stitching/last/pack), Factory, Order Received Date, AS XF.

Output:
  The cleaned workbook is saved to data/bronze/ by default, named
  preprocessed_<YYYY-MM-DD>.xlsx. An explicit output path can be provided
  as a second CLI argument.

Usage:
  uv run python -m skills.sort-shipment.scripts.preprocessor <input.xlsx> [output.xlsx]
  uv run python -m skills.sort-shipment.scripts.preprocessor "data/raw/AS_report_input.xlsx"
"""

from __future__ import annotations

from datetime import date, datetime
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet


class SheetCleaner:
    """Base class for worksheet cleaning operations."""

    HEADER_ROWS = 2

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws

    def find_columns_by_header(self, targets: list[str]) -> list[int]:
        """Return 1-based column indices whose header text matches any target (case-insensitive substring)."""
        matched: list[int] = []
        for row_idx in range(1, self.HEADER_ROWS + 1):
            for col_idx in range(1, self.ws.max_column + 1):
                val = self.ws.cell(row_idx, col_idx).value
                if val is None:
                    continue
                val_lower = str(val).strip().lower()
                for t in targets:
                    if t.lower() in val_lower and col_idx not in matched:
                        matched.append(col_idx)
        return sorted(matched)

    def unmerge_and_fill(self, col_indices: list[int]) -> None:
        """Unmerge cells in the given columns and propagate the top-left value."""
        target_cols = set(col_indices)
        for merged_range in list(self.ws.merged_cells.ranges):
            if merged_range.min_col not in target_cols:
                continue
            top_value = self.ws.cell(merged_range.min_row, merged_range.min_col).value
            self.ws.unmerge_cells(str(merged_range))
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    self.ws.cell(row, col).value = top_value

    def unmerge_header_rows(self) -> None:
        """Unmerge all merged cells in header rows and fill with top-left value."""
        for merged_range in list(self.ws.merged_cells.ranges):
            if merged_range.min_row > self.HEADER_ROWS:
                continue
            top_value = self.ws.cell(merged_range.min_row, merged_range.min_col).value
            self.ws.unmerge_cells(str(merged_range))
            for row in range(merged_range.min_row, merged_range.max_row + 1):
                for col in range(merged_range.min_col, merged_range.max_col + 1):
                    self.ws.cell(row, col).value = top_value

    def remove_images_in_columns(self, col_indices: list[int]) -> None:
        """Remove embedded images whose anchor falls within the given 1-based columns."""
        cols_0based = {c - 1 for c in col_indices}
        self.ws._images = [
            img for img in self.ws._images
            if not (hasattr(img, "anchor") and hasattr(img.anchor, "_from")
                    and img.anchor._from.col in cols_0based)
        ]

    def delete_columns(self, col_indices: list[int]) -> None:
        """Delete columns by 1-based index, right-to-left to avoid shifting."""
        self.remove_images_in_columns(col_indices)
        for col_idx in sorted(col_indices, reverse=True):
            self.ws.delete_cols(col_idx)

    def coerce_po_to_string(self) -> None:
        """Convert PO number cells from numeric to string so they are preserved as text."""
        po_cols = self.find_columns_by_header(["po"])
        if not po_cols:
            return
        for col_idx in po_cols:
            for row_idx in range(self.HEADER_ROWS + 1, self.ws.max_row + 1):
                cell = self.ws.cell(row_idx, col_idx)
                if cell.value is not None and not isinstance(cell.value, str):
                    cell.value = str(int(cell.value)) if isinstance(cell.value, float) and cell.value == int(cell.value) else str(cell.value)
                    cell.data_type = "s"

    def clean(self) -> None:
        """Override in subclasses to define sheet-specific cleaning steps."""
        raise NotImplementedError


class ASSheetCleaner(SheetCleaner):
    """Cleaner for the AS (pending shipments) worksheet."""

    REMOVE_HEADERS = [
        "photo", "wip", "cutting", "stitching", "last",
        "etd shenzhen", "as xf", "update as xf",
    ]

    def clean(self) -> None:
        self.unmerge_header_rows()
        self.coerce_po_to_string()
        cols_to_remove = self.find_columns_by_header(self.REMOVE_HEADERS)
        self.delete_columns(cols_to_remove)


class ShippedSheetCleaner(SheetCleaner):
    """Cleaner for the Shipped (completed shipments) worksheet."""

    FILL_HEADERS = ["lh xf", "container type", "etd-shenzhen", "eta-sa", "remark"]
    REMOVE_HEADERS = [
        "photo", "stitching", "last", "pack",
        "factory", "order received date", "as xf",
    ]

    def unhide_columns(self) -> None:
        for _col_letter, dim in self.ws.column_dimensions.items():
            if dim.hidden:
                dim.hidden = False

    def filter_rows_by_lh_xf_date(self) -> None:
        """Remove data rows where LH XF date is before January 1 of the current year."""
        lh_xf_cols = self.find_columns_by_header(["lh xf"])
        if not lh_xf_cols:
            return
        col_idx = lh_xf_cols[0]
        year_start = date(date.today().year, 1, 1)
        rows_to_delete: list[int] = []
        for row_idx in range(self.HEADER_ROWS + 1, self.ws.max_row + 1):
            val = self.ws.cell(row_idx, col_idx).value
            if val is None:
                continue
            if isinstance(val, datetime):
                val = val.date()
            if isinstance(val, date) and val < year_start:
                rows_to_delete.append(row_idx)
        for row_idx in reversed(rows_to_delete):
            self.ws.delete_rows(row_idx)

    def clean(self) -> None:
        self.unhide_columns()
        self.unmerge_header_rows()
        fill_cols = self.find_columns_by_header(self.FILL_HEADERS)
        self.unmerge_and_fill(fill_cols)
        self.coerce_po_to_string()
        self.filter_rows_by_lh_xf_date()
        cols_to_remove = self.find_columns_by_header(self.REMOVE_HEADERS)
        self.delete_columns(cols_to_remove)


class ShipmentPreprocessor:
    """Orchestrates loading, cleaning, and saving a vendor shipment Excel file."""

    SHEET_CLEANERS: list[type[SheetCleaner]] = [ASSheetCleaner, ShippedSheetCleaner]
    DEFAULT_OUTPUT_DIR = Path("data/bronze")

    def __init__(self, input_path: str | Path, output_path: str | Path | None = None) -> None:
        self.input_path = Path(input_path)
        if not self.input_path.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_path}")

        if output_path is None:
            self.DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            self.output_path = self.DEFAULT_OUTPUT_DIR / f"preprocessed_{date.today().isoformat()}.xlsx"
        else:
            self.output_path = Path(output_path)
            self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def run(self) -> Path:
        """Clean all sheets and save the preprocessed workbook."""
        wb = openpyxl.load_workbook(self.input_path)

        for idx, ws in enumerate(wb.worksheets):
            if idx < len(self.SHEET_CLEANERS):
                cleaner = self.SHEET_CLEANERS[idx](ws)
                cleaner.clean()

        wb.save(self.output_path)
        print(f"Preprocessed file saved to: {self.output_path}")
        return self.output_path


if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: uv run python -m skills.sort-shipment.scripts.preprocessor <input.xlsx> [output.xlsx]")
        sys.exit(1)

    in_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else None
    preprocessor = ShipmentPreprocessor(in_file, out_file)
    preprocessor.run()
