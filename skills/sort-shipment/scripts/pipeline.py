"""
Deterministic shipment extraction pipeline.

Replaces the LLM-based extraction step with rule-based parsing.
Reads preprocessed vendor Excel data and produces structured shipment
records following the SKILL.md output schema.

Usage:
    uv run python -m skills.sort-shipment.scripts.pipeline data/bronze/preprocessed_2026-04-04.xlsx
"""

from __future__ import annotations

import logging
import re
import shutil
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl

from .models import Shipment

if TYPE_CHECKING:
    from openpyxl.worksheet.worksheet import Worksheet

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Date normalisation
# ---------------------------------------------------------------------------

class DateNormalizer:
    """Parse messy date strings into ``date`` objects.

    Handles formats commonly found in vendor Excel files:
    - YYYY/MM/DD, YYYY-MM-DD
    - MM/DD, MM-DD  (year inferred from context)
    - datetime objects from openpyxl
    """

    # Order matters – try most specific patterns first.
    _PATTERNS: list[tuple[re.Pattern[str], str]] = [
        # Allow optional whitespace around separators (vendors write "2026/ 6/8").
        (re.compile(r"(\d{4})\s*[/-]\s*(\d{1,2})\s*[/-]\s*(\d{1,2})"), "ymd"),
        (re.compile(r"(\d{1,2})\s*[/-]\s*(\d{1,2})"), "md"),
    ]

    def __init__(self, default_year: int | None = None) -> None:
        self.default_year = default_year or date.today().year

    def parse(self, value: object) -> date | None:
        """Return a ``date`` or *None* if the value cannot be resolved."""
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value

        if value is None:
            return None

        text = str(value).strip()
        if not text:
            return None

        for pattern, kind in self._PATTERNS:
            m = pattern.search(text)
            if m:
                try:
                    if kind == "ymd":
                        return date(int(m.group(1)), int(m.group(2)), int(m.group(3)))
                    # kind == "md"
                    return date(self.default_year, int(m.group(1)), int(m.group(2)))
                except ValueError:
                    continue
        return None

    def parse_first_date(self, text: str) -> date | None:
        """Extract the *first* recognisable date from arbitrary text."""
        return self.parse(text)


# ---------------------------------------------------------------------------
# Shipment split detection
# ---------------------------------------------------------------------------

class ShipmentSplitter:
    """Detect and resolve multi-shipment rows (SKILL.md Steps 1A & 1B).

    A single row may encode several shipments when the LH XF cell contains
    multiple ``date:qty`` lines.
    """

    # Matches lines like "4/1:660", "2026/4/1: 1314", "3/26:600 OK"
    _LINE_RE = re.compile(
        r"(?P<date>[\d/\-]+)"       # date portion
        r"\s*[:：]\s*"               # colon separator
        r"(?P<qty>\d+)"             # quantity
        r"(?P<suffix>.*)",          # trailing text (e.g. " OK", " Air OK")
    )

    _OK_RE = re.compile(r"\bOK\b", re.IGNORECASE)

    def split(self, pairs: int, lh_xf_raw: str, eta_sa_raw: str | None) -> list[dict]:
        """Return a list of dicts with keys ``date``, ``qty``, ``eta_sa_raw``.

        Each dict represents one shipment split.  If no split is detected the
        list contains a single entry with the original values.
        """
        if not lh_xf_raw or not lh_xf_raw.strip():
            return [{"date_raw": None, "qty": pairs, "eta_sa_raw": eta_sa_raw}]

        lines = [ln.strip() for ln in str(lh_xf_raw).splitlines() if ln.strip()]
        parsed = self._parse_lines(lines)

        if not parsed:
            # No structured lines found – treat the whole cell as a single date.
            return [{"date_raw": lh_xf_raw.strip(), "qty": pairs, "eta_sa_raw": eta_sa_raw}]

        # Separate completed ("OK") and pending shipments.
        pending = [p for p in parsed if not p["ok"]]
        completed = [p for p in parsed if p["ok"]]

        # --- Case B: partial shipments already completed ---
        # If there are completed lines, find the pending line whose qty
        # matches total pairs.
        if completed and pending:
            for p in pending:
                if p["qty"] == pairs:
                    return [self._to_result(p, eta_sa_raw, lines, parsed)]
            # Fallback: return all pending.

        # --- Case A: multi-line XF with matching total ---
        if len(pending) > 1:
            total = sum(p["qty"] for p in pending)
            if total == pairs:
                return self._split_with_eta(pending, eta_sa_raw, lines)

        # Single pending line or unresolvable – return pending lines as-is.
        if pending:
            return self._split_with_eta(pending, eta_sa_raw, lines)

        # Everything is completed – return nothing pending.
        return []

    # -- helpers --

    def _parse_lines(self, lines: list[str]) -> list[dict]:
        results = []
        for line in lines:
            m = self._LINE_RE.match(line)
            if m:
                results.append({
                    "date_raw": m.group("date"),
                    "qty": int(m.group("qty")),
                    "ok": bool(self._OK_RE.search(m.group("suffix"))),
                    "suffix": m.group("suffix").strip(),
                })
        return results

    def _to_result(self, parsed: dict, eta_sa_raw: str | None,
                   lines: list[str], all_parsed: list[dict]) -> dict:
        """Convert a parsed line dict to the output format."""
        eta = self._match_eta(parsed, eta_sa_raw, lines, all_parsed)
        return {"date_raw": parsed["date_raw"], "qty": parsed["qty"], "eta_sa_raw": eta}

    def _split_with_eta(self, pending: list[dict], eta_sa_raw: str | None,
                        lines: list[str]) -> list[dict]:
        """Distribute ETA-SA lines across pending shipments."""
        eta_lines = self._split_eta_lines(eta_sa_raw)
        results = []
        for i, p in enumerate(pending):
            eta = eta_lines[i] if i < len(eta_lines) else eta_sa_raw
            results.append({"date_raw": p["date_raw"], "qty": p["qty"], "eta_sa_raw": eta})
        return results

    @staticmethod
    def _split_eta_lines(eta_sa_raw: str | None) -> list[str]:
        if not eta_sa_raw:
            return []
        return [ln.strip() for ln in str(eta_sa_raw).splitlines() if ln.strip()]

    @staticmethod
    def _match_eta(parsed: dict, eta_sa_raw: str | None,
                   lines: list[str], all_parsed: list[dict]) -> str | None:
        """Try to pick the ETA-SA line that corresponds to this shipment."""
        eta_lines = ShipmentSplitter._split_eta_lines(eta_sa_raw)
        if not eta_lines:
            return eta_sa_raw
        # Find the index of this parsed entry among *all* parsed entries.
        idx = next((i for i, p in enumerate(all_parsed) if p is parsed), 0)
        # Pending-only index among non-OK entries.
        pending_idx = sum(1 for p in all_parsed[:idx] if not p.get("ok", False))
        if pending_idx < len(eta_lines):
            return eta_lines[pending_idx]
        return eta_sa_raw


# ---------------------------------------------------------------------------
# Derived field computation
# ---------------------------------------------------------------------------

class DerivedFieldCalculator:
    """Compute ``eta_sa`` and ``eta_fac`` from origin and ``lh_xf``."""

    _CAMBODIA_RE = re.compile(r"cambodia|柬埔寨|柬國", re.IGNORECASE)
    CAMBODIA_TRANSIT_DAYS = 55
    DEFAULT_TRANSIT_DAYS = 35
    ETA_FAC_OFFSET_DAYS = 3

    def is_cambodia(self, remark: str | None) -> bool:
        if not remark:
            return False
        return bool(self._CAMBODIA_RE.search(remark))

    def compute_eta_sa(self, lh_xf: date | None, remark: str | None,
                       raw_eta_sa: date | None) -> date | None:
        """Return ``eta_sa``: use the raw value if it is a real date,
        otherwise compute from ``lh_xf`` + transit days."""
        if raw_eta_sa is not None:
            return raw_eta_sa
        if lh_xf is None:
            return None
        days = self.CAMBODIA_TRANSIT_DAYS if self.is_cambodia(remark) else self.DEFAULT_TRANSIT_DAYS
        return lh_xf + timedelta(days=days)

    def compute_eta_fac(self, eta_sa: date | None) -> date | None:
        if eta_sa is None:
            return None
        return eta_sa + timedelta(days=self.ETA_FAC_OFFSET_DAYS)


# ---------------------------------------------------------------------------
# Customer-requested XF parser
# ---------------------------------------------------------------------------

class CustomerRequestedXFParser:
    """Extract customer-requested ex-factory date from Remark text.

    Matches patterns like:
    - "Customer Required XF 2026/6/19"
    - "Cusomter request XF at 2026/5/1"   (typo-tolerant)
    - "客人要求4/15出"
    - "cust req XF 4/15"
    """

    # Date fragment allowing optional spaces around separators.
    _DATE_FRAG = r"\d{4}\s*[/-]\s*\d{1,2}\s*[/-]\s*\d{1,2}|\d{1,2}\s*[/-]\s*\d{1,2}"

    _PATTERNS: list[re.Pattern[str]] = [
        # English variants – very tolerant of typos ("Cusomter", "Custmer", etc.)
        re.compile(
            r"[Cc]us\w*(?:er|or|re|te)?\s+[Rr]eq\w*\s+XF\s*(?:at\s+)?"
            rf"(?P<date>{_DATE_FRAG})",
            re.IGNORECASE,
        ),
        # Chinese: 客人要求 <date> 出
        re.compile(
            rf"客人要求\s*(?P<date>{_DATE_FRAG})\s*出?",
        ),
        # Short form: "cust req XF 4/15"
        re.compile(
            rf"cust\s+req\s+XF\s+(?P<date>{_DATE_FRAG})",
            re.IGNORECASE,
        ),
    ]

    def __init__(self, date_normalizer: DateNormalizer) -> None:
        self._dn = date_normalizer

    def parse(self, remark: str | None) -> date | None:
        if not remark:
            return None
        for pattern in self._PATTERNS:
            m = pattern.search(remark)
            if m:
                return self._dn.parse(m.group("date"))
        return None


# ---------------------------------------------------------------------------
# Sheet reader
# ---------------------------------------------------------------------------

class SheetReader:
    """Read a preprocessed worksheet into a list of raw row dicts.

    Handles column mapping for both AS and Shipped sheet layouts.
    """

    HEADER_ROWS = 2

    # Canonical column names → possible header substrings (case-insensitive).
    _COLUMN_MAP: dict[str, list[str]] = {
        "po_number": ["customer po", "po#", "po #"],
        "brand": ["brand"],
        "style": ["style"],
        "pairs": ["pair"],
        "lh_xf": ["lh", "xf"],
        "eta_sa": ["eta-sa", "eta sa", "est.eta"],
        "remark": ["remark"],
        "container_type": ["container type"],
        "etd_port": ["etd", "shenzhen"],
        "container_number": ["container number", "container no"],
    }

    def __init__(self, ws: Worksheet) -> None:
        self.ws = ws
        self._col_indices: dict[str, int] = {}
        self._resolve_columns()

    def _resolve_columns(self) -> None:
        """Map canonical field names to 1-based column indices."""
        # Build a header text map: col_idx → combined header text from rows 1-2.
        header_texts: dict[int, str] = {}
        for col_idx in range(1, self.ws.max_column + 1):
            parts = []
            for row_idx in range(1, self.HEADER_ROWS + 1):
                val = self.ws.cell(row_idx, col_idx).value
                if val:
                    parts.append(str(val).strip().lower())
            header_texts[col_idx] = " ".join(parts)

        for field, candidates in self._COLUMN_MAP.items():
            for col_idx, header in header_texts.items():
                if any(c in header for c in candidates):
                    # Avoid double-mapping: prefer first match unless already taken.
                    if field not in self._col_indices:
                        self._col_indices[field] = col_idx
                        break

    def read_rows(self) -> list[dict[str, object]]:
        """Return data rows as a list of dicts keyed by canonical field name."""
        rows: list[dict[str, object]] = []
        for row_idx in range(self.HEADER_ROWS + 1, self.ws.max_row + 1):
            row: dict[str, object] = {}
            all_empty = True
            for field, col_idx in self._col_indices.items():
                val = self.ws.cell(row_idx, col_idx).value
                row[field] = val
                if val is not None and str(val).strip():
                    all_empty = False
            if not all_empty:
                rows.append(row)
        return rows


# ---------------------------------------------------------------------------
# ETA-SA parser for Shipped sheet (handles "ETA LAX: 1/31\nETA SA: 2/05")
# ---------------------------------------------------------------------------

class ETASAParser:
    """Extract the SA-specific ETA from multi-line ETA cells."""

    _SA_RE = re.compile(
        r"ETA\s+SA\s*[:：]\s*(?P<date>\d{4}[/-]\d{1,2}[/-]\d{1,2}|\d{1,2}[/-]\d{1,2})",
        re.IGNORECASE,
    )
    _ETA_RE = re.compile(
        r"ETA\s*[:：]\s*(?P<date>\d{4}[/-]\d{1,2}[/-]\d{1,2}|\d{1,2}[/-]\d{1,2})",
        re.IGNORECASE,
    )

    def __init__(self, date_normalizer: DateNormalizer) -> None:
        self._dn = date_normalizer

    def parse(self, raw: object) -> date | None:
        """Return the best ETA-SA date from the raw cell value."""
        # If it's already a date/datetime from openpyxl, use directly.
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw

        if raw is None:
            return None

        text = str(raw).strip()
        if not text:
            return None

        # Prefer "ETA SA:" line over generic "ETA:".
        m = self._SA_RE.search(text)
        if m:
            return self._dn.parse(m.group("date"))

        # If the cell contains only a single date (no "ETA SA" label), parse it.
        # But skip if it looks like "ETA LAX:" only.
        if "ETA" in text.upper() and "SA" not in text.upper():
            # Only LAX or other port – fall back to generic date parse.
            pass

        return self._dn.parse(text)


# ---------------------------------------------------------------------------
# ETD port parser (handles "ETD: 1/16", "ETD:1/17", or plain dates)
# ---------------------------------------------------------------------------

class ETDPortParser:
    """Extract ETD port date from cell values."""

    _ETD_RE = re.compile(
        r"ETD\s*[:：]\s*(?P<date>\d{4}[/-]\d{1,2}[/-]\d{1,2}|\d{1,2}[/-]\d{1,2})",
        re.IGNORECASE,
    )

    def __init__(self, date_normalizer: DateNormalizer) -> None:
        self._dn = date_normalizer

    def parse(self, raw: object) -> date | None:
        if isinstance(raw, datetime):
            return raw.date()
        if isinstance(raw, date):
            return raw
        if raw is None:
            return None

        text = str(raw).strip()
        m = self._ETD_RE.search(text)
        if m:
            return self._dn.parse(m.group("date"))
        return self._dn.parse(text)


# ---------------------------------------------------------------------------
# Core parser: raw rows → Shipment records
# ---------------------------------------------------------------------------

class ShipmentParser:
    """Transform raw row dicts into validated ``Shipment`` records."""

    def __init__(self) -> None:
        self._dn = DateNormalizer()
        self._splitter = ShipmentSplitter()
        self._derived = DerivedFieldCalculator()
        self._cust_xf = CustomerRequestedXFParser(self._dn)
        self._eta_sa_parser = ETASAParser(self._dn)
        self._etd_parser = ETDPortParser(self._dn)

    def parse_rows(self, rows: list[dict[str, object]]) -> list[Shipment]:
        """Parse all rows, returning valid ``Shipment`` objects.

        Rows that cannot be resolved are skipped with a warning.
        """
        shipments: list[Shipment] = []
        # Track shipment_idx per PO.
        po_idx_counter: dict[str, int] = {}

        for row in rows:
            try:
                parsed = self._parse_single_row(row, po_idx_counter)
                shipments.extend(parsed)
            except Exception:
                po = row.get("po_number", "???")
                logger.warning("Skipping unresolvable row for PO %s", po, exc_info=True)
        return shipments

    def _parse_single_row(self, row: dict[str, object],
                          po_idx: dict[str, int]) -> list[Shipment]:
        po_number = self._clean_po(row.get("po_number"))
        if not po_number:
            logger.warning("Skipping row with missing PO number")
            return []

        pairs_raw = row.get("pairs")
        pairs = self._parse_pairs(pairs_raw)
        if pairs is None:
            logger.warning("Skipping PO %s: cannot determine pairs from %r", po_number, pairs_raw)
            return []

        brand = str(row.get("brand", "") or "").strip()
        style = str(row.get("style", "") or "").strip()
        remark = str(row.get("remark", "") or "").strip() or None

        lh_xf_raw = row.get("lh_xf")
        eta_sa_raw = row.get("eta_sa")

        # Container fields.
        container_type = str(row.get("container_type", "") or "").strip() or None
        container_number = str(row.get("container_number", "") or "").strip() or None

        # ETD port.
        etd_port = self._etd_parser.parse(row.get("etd_port"))

        # Customer-requested XF from remark.
        customer_requested_xf = self._cust_xf.parse(remark)

        # Detect shipment splits.
        lh_xf_text = str(lh_xf_raw) if lh_xf_raw is not None else ""
        eta_sa_text = str(eta_sa_raw) if eta_sa_raw is not None else ""

        # If lh_xf is already a date object (not multi-line), skip splitting.
        if isinstance(lh_xf_raw, (date, datetime)):
            splits = [{"date_raw": lh_xf_raw, "qty": pairs, "eta_sa_raw": eta_sa_raw}]
        else:
            splits = self._splitter.split(pairs, lh_xf_text, eta_sa_text)

        if not splits:
            logger.warning("Skipping PO %s: no resolvable shipments", po_number)
            return []

        results: list[Shipment] = []
        for split in splits:
            idx = po_idx.get(po_number, 0) + 1
            po_idx[po_number] = idx

            lh_xf = self._dn.parse(split["date_raw"])
            raw_eta_sa = self._eta_sa_parser.parse(split.get("eta_sa_raw"))
            eta_sa = self._derived.compute_eta_sa(lh_xf, remark, raw_eta_sa)
            eta_fac = self._derived.compute_eta_fac(eta_sa)

            results.append(Shipment(
                po_number=po_number,
                shipment_idx=idx,
                brand=brand,
                style=style,
                pairs=split["qty"],
                lh_xf=lh_xf,
                etd_port=etd_port,
                eta_sa=eta_sa,
                eta_fac=eta_fac,
                customer_requested_xf=customer_requested_xf,
                container_type=container_type or "",
                container_number=container_number,
                remark=remark,
            ))
        return results

    # -- helpers --

    @staticmethod
    def _clean_po(value: object) -> str | None:
        if value is None:
            return None
        s = str(value).strip()
        if not s:
            return None
        # Remove trailing ".0" from float-cast POs.
        if s.endswith(".0"):
            s = s[:-2]
        return s

    @staticmethod
    def _parse_pairs(value: object) -> int | None:
        if value is None:
            return None
        if isinstance(value, (int, float)):
            return int(value)
        text = str(value).strip().replace(",", "")
        m = re.search(r"\d+", text)
        if m:
            return int(m.group())
        return None


# ---------------------------------------------------------------------------
# Gantt writer
# ---------------------------------------------------------------------------

class GanttWriter:
    """Write ``Shipment`` records into the gantt template .xlsm file."""

    TEMPLATE_PATH = Path("skills/sort-shipment/references/gantt_template.xlsm")
    DEFAULT_OUTPUT_DIR = Path("data/silver")

    def __init__(self, output_path: str | Path | None = None) -> None:
        if output_path is None:
            self.DEFAULT_OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
            self.output_path = self.DEFAULT_OUTPUT_DIR / f"gantt_{date.today().isoformat()}.xlsm"
        else:
            self.output_path = Path(output_path)
            self.output_path.parent.mkdir(parents=True, exist_ok=True)

    def write(self, shipments: list[Shipment]) -> Path:
        """Copy template and write shipment data starting at ``data_dest``."""
        shutil.copy2(self.TEMPLATE_PATH, self.output_path)

        wb = openpyxl.load_workbook(self.output_path, keep_vba=True)
        ws = wb["data"]

        # data_dest is defined as data!$A$2 → start writing at row 2.
        start_row = 2

        # Clear any existing data rows (from row 2 onward).
        for row_idx in range(start_row, ws.max_row + 1):
            for col_idx in range(1, 14):
                ws.cell(row_idx, col_idx).value = None

        for i, s in enumerate(shipments):
            row = start_row + i
            ws.cell(row, 1).value = s.po_number
            ws.cell(row, 1).data_type = "s"  # Keep PO as string.
            ws.cell(row, 2).value = s.shipment_idx
            ws.cell(row, 3).value = s.brand
            ws.cell(row, 4).value = s.style
            ws.cell(row, 5).value = s.pairs
            ws.cell(row, 6).value = s.lh_xf
            ws.cell(row, 7).value = s.etd_port
            ws.cell(row, 8).value = s.eta_sa
            ws.cell(row, 9).value = s.eta_fac
            ws.cell(row, 10).value = s.customer_requested_xf
            ws.cell(row, 11).value = s.container_type
            ws.cell(row, 12).value = s.container_number
            ws.cell(row, 13).value = s.remark

        wb.save(self.output_path)
        return self.output_path


# ---------------------------------------------------------------------------
# Pipeline orchestrator
# ---------------------------------------------------------------------------

class SortShipmentPipeline:
    """End-to-end orchestrator: read → parse → write."""

    def __init__(self, input_path: str | Path,
                 output_path: str | Path | None = None) -> None:
        self.input_path = Path(input_path)
        if not self.input_path.exists():
            raise FileNotFoundError(f"Input file not found: {self.input_path}")
        self.output_path = output_path

    def run(self) -> Path:
        """Execute the full pipeline and return the output file path."""
        logging.basicConfig(
            level=logging.INFO,
            format="%(levelname)s: %(message)s",
        )

        wb = openpyxl.load_workbook(self.input_path, data_only=True)
        parser = ShipmentParser()
        all_shipments: list[Shipment] = []

        for ws in wb.worksheets:
            logger.info("Reading sheet: %s", ws.title)
            reader = SheetReader(ws)
            rows = reader.read_rows()
            logger.info("  Found %d data rows", len(rows))
            shipments = parser.parse_rows(rows)
            logger.info("  Extracted %d shipment records", len(shipments))
            all_shipments.extend(shipments)

        logger.info("Total shipments: %d", len(all_shipments))

        writer = GanttWriter(self.output_path)
        out = writer.write(all_shipments)
        logger.info("Output written to: %s", out)
        return out


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

if __name__ == "__main__":
    import sys

    if len(sys.argv) < 2:
        print("Usage: uv run python -m skills.sort-shipment.scripts.pipeline <preprocessed.xlsx> [output.xlsm]")
        sys.exit(1)

    in_file = sys.argv[1]
    out_file = sys.argv[2] if len(sys.argv) > 2 else None
    pipeline = SortShipmentPipeline(in_file, out_file)
    pipeline.run()
