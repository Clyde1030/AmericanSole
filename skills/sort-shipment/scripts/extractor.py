"""
Excel Extractor — reads a messy vendor Excel file and converts it to
flat text that Claude Code can reason about.
"""

from pathlib import Path

import openpyxl


def _sheet_to_text(ws) -> str:
    """
    Convert a single worksheet to a plain-text representation.

    Handles merged cells by writing the value into every cell of the
    merged range so the full context is visible without gaps.
    """
    merge_map: dict[tuple[int, int], object] = {}
    for merged_range in list(ws.merged_cells.ranges):
        top_left = ws.cell(merged_range.min_row, merged_range.min_col)
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                merge_map[(row, col)] = top_left.value

    rows_text: list[str] = []
    for row in ws.iter_rows():
        cells: list[str] = []
        for cell in row:
            value = merge_map.get((cell.row, cell.column), cell.value)
            cells.append("" if value is None else str(value).strip().replace("\n", " | "))
        if any(c for c in cells):
            rows_text.append("\t".join(cells))

    return "\n".join(rows_text)


def excel_to_text(file_path: str | Path) -> list[tuple[str, str]]:
    """
    Read a vendor Excel file and return each sheet as plain text.

    Args:
        file_path: Path to the .xlsx file.

    Returns:
        List of (sheet_name, sheet_text) for every non-empty sheet.
    """
    path = Path(file_path)
    if not path.exists():
        raise FileNotFoundError(f"Input file not found: {path}")

    wb = openpyxl.load_workbook(path, data_only=True)
    result = []
    for sheet_name in wb.sheetnames:
        text = _sheet_to_text(wb[sheet_name])
        if text.strip():
            result.append((sheet_name, text))
    return result
